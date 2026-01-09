#!/usr/bin/env python3

"""
Excel Sheet Unlocker
Author: (r0gg)
Description: Remove sheet protection from Excel files
"""

__version__ = "1.0.0"
__author__ = "(r0gg)"


def print_banner():
    """Display tool banner"""
    banner = """
╔═══════════════════════════════════════════════════════════╗
║                                                           ║
║        🔓 EXCEL SHEET UNLOCKER v{}                    ║
║                                                           ║
║        Author: {}                                  ║
║        Purpose: Remove Excel sheet protection            ║
║                                                           ║
╚═══════════════════════════════════════════════════════════╝
    """.format(__version__, __author__)
    print(banner)

def main():
    """Main function"""
    print_banner()


if __name__ == "__main__":
    main()

from openpyxl import load_workbook
from openpyxl.styles import Protection
import sys

wb = load_workbook(sys.argv[1])
unlocked = Protection(locked=False)

for sheet in wb.worksheets:
    sheet.protection.sheet = False
    for row in sheet.iter_rows():
        for cell in row:
            cell.protection = unlocked
    print(f"[+] {sheet.title} débloquée")

output = sys.argv[1].replace('.xlsx', '_UNLOCKED.xlsx')
wb.save(output)
print(f"[✓] Sauvegardé: {output}")


